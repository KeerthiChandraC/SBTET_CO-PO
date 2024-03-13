from openpyxl import load_workbook
from openpyxl.styles import Border, Side

OV_DATA='BACDEFGHIJK'
OV_DATA_L='BCDEFGHIJ'

# def copy_worksheet(source_ws, target_ws):
#     for row in source_ws.iter_rows():
#         for cell in row:
#             target_ws[cell.coordinate].value = cell.value
#             target_ws[cell.coordinate].font = cell.font.copy()
#             target_ws[cell.coordinate].border = cell.border.copy()
#             target_ws[cell.coordinate].fill = cell.fill.copy()
#             target_ws[cell.coordinate].number_format = cell.number_format
#             target_ws[cell.coordinate].protection = cell.protection.copy()
#             target_ws[cell.coordinate].alignment = cell.alignment.copy()
#             target_ws[cell.coordinate].comment = cell.comment

def main():
    temp_wb = load_workbook('Final template.xlsx')
    temp_ws = temp_wb['temp']
    source_wb = load_workbook('CO-PO Matrix Gen.xlsx')
    source_ws = source_wb['REF']
    source_ovrl = source_wb['OVERALL']
    for row_idx in range(1, temp_ws.max_row+1):
         
        target_ws = source_wb.copy_worksheet(source_ws)
        

        target_ws.title = f'{temp_ws.cell(row=row_idx, column=1).value}'

        target_ws['B2'] = f'{temp_ws.cell(row=row_idx, column=1).value}'
        target_ws['B1'] = f'{temp_ws.cell(row=row_idx, column=2).value}'
        for sr in range(1, len(OV_DATA)):
            source_ovrl.cell(row=row_idx, column=sr).value = f"='{temp_ws.cell(row=row_idx, column=1).value}'!{OV_DATA[sr]}19"
            
        source_wb.move_sheet(target_ws, -1)
    for sr in range(1, len(OV_DATA_L)+1):
        source_ovrl.cell(row=row_idx+1, column=sr+1).value =  f'=IFERROR(AVERAGE({OV_DATA_L[sr-1]}1:{OV_DATA_L[sr-1]}{row_idx}),"-")'
        
    source_ovrl.cell(row=row_idx+1, column=1).value =  f'Total Average:'
    thin_border = Side(border_style='thin', color='000000')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    ranges = source_ovrl[f"A1:{OV_DATA_L[sr-1]}{row_idx+1}"] 
    for cell in ranges:
        for x in cell:
            x.border=border
    source_wb.save('new.xlsx')
    
if __name__ == "__main__":
    main()
    print("Completed")